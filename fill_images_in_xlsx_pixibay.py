#!/usr/bin/env python3
"""
fill_images_in_xlsx_pixabay_multi.py
------------------------------------

WHAT THIS DOES
--------------
- Opens your Excel workbook that contains *multiple sheets*.
- On every sheet, it looks for a column named "Title" (product title).
- It makes a **Pixabay API** search (CC0-friendly images) for each missing image.
- When it finds a result, it writes the image URL into the "Image" column.
- It saves a new workbook with the Image column filled.

WHY PIXABAY?
------------
- Pixabay images are free to use (CC0-like license). Always review their terms:
  https://pixabay.com/service/terms/
- Using an API is more reliable than scraping Google/Bing, and safer for ToS.

REQUIREMENTS (one-time)
-----------------------
1) Python 3.12 (you have it) + a virtual environment is recommended:

   python3 -m venv ~/woo-import-images/.venv
   source ~/woo-import-images/.venv/bin/activate
   pip install openpyxl requests

2) Get a free Pixabay API key:
   https://pixabay.com/api/docs/

3) Tell the script your API key (in the same terminal session):
   export PIXABAY_API_KEY=YOUR_KEY_HERE

HOW TO RUN
----------
- Put this script next to your workbook:
    products_for_auto_images_ALL_SHEETS.xlsx
- Run:
    python3 fill_images_in_xlsx_pixabay_multi.py

OUTPUT
------
- A new file named:
    products_with_images_ALL_SHEETS_PIXABAY.xlsx

TROUBLESHOOTING
---------------
- If you get "PIXABAY_API_KEY is not set", export it again in this terminal.
- If some products still have no images, Pixabay likely had no good match for
  that query. You can adjust the query building below to include Brand/Region.

SAFE LIMITS / USAGE
-------------------
- The script delays between queries to respect API limits (see DELAY below).
- It also does a HEAD request to confirm the image URL is alive before saving.
"""

import os
import time
import requests
from typing import Optional
from openpyxl import load_workbook

# ====================== USER CONFIG (EDIT THESE) ======================

# Name of your input workbook (the one we prepared with Title + Image columns).
INPUT_XLSX  = "products_for_auto_images_ALL_SHEETS.xlsx"

# Name of the output workbook that will be created.
OUTPUT_XLSX = "products_with_images_ALL_SHEETS_PIXABAY.xlsx"

# Column names in your sheets:
TITLE_HEADER = "Title"   # Where we read the product title/search phrase from
IMAGE_HEADER = "Image"   # Where we write the found image URL

# Networking / throttling:
REQUEST_TIMEOUT = 20     # Seconds to wait for each web request
DELAY_BETWEEN_QUERIES = 0.6   # Small delay (seconds) between API calls

# Search preferences for Pixabay
MIN_WIDTH  = 600            # prefer medium+ images
MIN_HEIGHT = 600
SAFESEARCH = "true"         # "true" or "false"

# If your sheet also has other columns, you can use them to build a richer query.
# Example: include "Brand" or "Region" to help the search.
# The script will try to find these columns; if missing, it ignores them.
OPTIONAL_QUERY_COLUMNS = ["Brand", "Region"]

# ================== END OF USER CONFIG (USUALLY OK) ===================

# Read API key from environment variable; this is safer than hard-coding it.
API_KEY = os.getenv("PIXABAY_API_KEY", "").strip()
API_URL = "https://pixabay.com/api/"


def head_ok(url: str, timeout: int = 10) -> bool:
    """
    Quick check that the URL responds with a 2xx or 3xx code before we save it.
    This avoids storing dead links in your Excel.
    """
    try:
        r = requests.head(url, timeout=timeout, allow_redirects=True)
        return 200 <= r.status_code < 400
    except Exception:
        return False


def build_query(base_title: str, ws, row_idx: int, col_index_by_name: dict) -> str:
    """
    Build a search query for Pixabay. Start with the Title, and optionally
    add more context (e.g., Brand/Region) if those columns exist.

    You can customize this function freely to improve matching.
    """
    parts = [base_title.strip()]

    # If Brand/Region columns exist, append them
    for col_name in OPTIONAL_QUERY_COLUMNS:
        col_idx = col_index_by_name.get(col_name)
        if col_idx:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and str(val).strip():
                parts.append(str(val).strip())

    # Example: add a generic word like "gift card" if you want to bias results
    # parts.append("gift card")

    # Final query string
    return " ".join(parts)


def pixabay_first_image_url(query: str) -> Optional[str]:
    """
    Query the Pixabay API for a given search phrase, and return the best image URL
    (prefer largeImageURL, then webformatURL, then previewURL). Returns None if
    nothing reasonable is found.

    Docs: https://pixabay.com/api/docs/
    """
    if not API_KEY:
        # Fail fast with a clear message so the user knows what to do.
        raise RuntimeError("PIXABAY_API_KEY is not set. Export it in your shell: "
                           "export PIXABAY_API_KEY=YOUR_KEY_HERE")

    params = {
        "key": API_KEY,
        "q": query,                 # the search string
        "image_type": "photo",      # photos only; you can use 'illustration' if needed
        "safesearch": SAFESEARCH,   # filter adult content
        "per_page": 10,             # how many results to fetch per request
        "page": 1,
        "lang": "en",               # language for searching
        "orientation": "horizontal",
        "order": "popular",
        "min_width": MIN_WIDTH,
        "min_height": MIN_HEIGHT,
    }

    try:
        r = requests.get(API_URL, params=params, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
    except Exception:
        return None

    data = r.json()
    hits = data.get("hits", [])
    if not hits:
        return None

    # Prefer the larger image when available. Verify the URL is alive.
    for h in hits:
        candidate = h.get("largeImageURL") or h.get("webformatURL") or h.get("previewURL")
        if not candidate:
            continue
        if head_ok(candidate, timeout=10):
            return candidate

    return None


def ensure_headers(ws, headers_needed: list[str]) -> dict[str, int]:
    """
    Ensure required headers exist in row 1. If a header is missing, create it.
    Return a mapping {header_name: column_index} for easy access later.
    """
    header_row = 1
    existing = {}
    max_col = ws.max_column or 1

    # Read current headers from row 1 (A1 -> first cell)
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str) and val.strip():
            existing[val.strip()] = col

    # Add any missing headers to the right
    for h in headers_needed:
        if h not in existing:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=h)
            existing[h] = max_col

    return existing


def main():
    # Basic sanity check
    if not os.path.exists(INPUT_XLSX):
        print(f"[!] Input workbook not found: {INPUT_XLSX}")
        return

    # Load workbook once; we’ll iterate all sheets
    wb = load_workbook(INPUT_XLSX)

    total_filled = 0
    total_skipped = 0

    # Process every sheet in the workbook
    for ws in wb.worksheets:
        # Make sure the two columns we depend on exist
        col_map = ensure_headers(ws, [TITLE_HEADER, IMAGE_HEADER])

        # Build a handy "name -> index" map we can also use for optional columns
        col_index_by_name = {k: v for k, v in col_map.items()}

        title_col = col_map[TITLE_HEADER]
        image_col = col_map[IMAGE_HEADER]

        rows = ws.max_row
        filled = 0
        skipped = 0

        print(f"\nSheet: {ws.title} — rows: {rows - 1}")
        print(f"Filling '{IMAGE_HEADER}' from '{TITLE_HEADER}' using Pixabay...")

        # Start from row 2 (row 1 is headers)
        for row in range(2, rows + 1):
            title_val = ws.cell(row=row, column=title_col).value

            # Skip rows that don’t have a title
            if not title_val or not str(title_val).strip():
                skipped += 1
                continue

            # If Image already has a value, don’t overwrite it
            existing_image = ws.cell(row=row, column=image_col).value
            if existing_image and str(existing_image).strip():
                continue

            # Build a richer query string (Title + optional columns)
            query = build_query(str(title_val), ws, row, col_index_by_name)
            print(f"- [{ws.title} R{row}] Pixabay search: {query!r}")

            url = pixabay_first_image_url(query)

            if url:
                ws.cell(row=row, column=image_col, value=url)
                filled += 1
            else:
                print("  (no CC0 image found)")

            # Be polite to the API: a small pause between requests
            time.sleep(DELAY_BETWEEN_QUERIES)

        print(f"Sheet '{ws.title}': filled {filled}, skipped {skipped}")
        total_filled += filled
        total_skipped += skipped

    # Save the new workbook with URLs filled in
    wb.save(OUTPUT_XLSX)
    print("\nAll done.")
    print(f"Total filled:  {total_filled}")
    print(f"Total skipped: {total_skipped}")
    print(f"Saved as:      {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
