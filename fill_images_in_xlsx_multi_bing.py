#!/usr/bin/env python3
"""
fill_images_in_xlsx_multi_bing.py
---------------------------------

WHAT THIS DOES
--------------
- Opens an Excel workbook with MULTIPLE SHEETS.
- Ensures each sheet has "Title" and "Image" headers.
- For each row without an Image, searches Bing Images using the Title
  (plus optional Brand/Region), finds the first *live* image URL, and writes it.
- Saves a new workbook with the Image column filled.

NO API KEYS NEEDED (uses Bing web results).

PROS / CONS
-----------
+ Quick to set up, wide coverage.
- Not license-filtered: use responsibly, or restrict to allowed domains below.

REQUIREMENTS
------------
1) Python 3.12 and a virtual environment (recommended):
   python3 -m venv ~/woo-import-images/.venv
   source ~/woo-import-images/.venv/bin/activate
   pip install openpyxl requests beautifulsoup4 lxml

2) Put this script next to your workbook:
   products_for_auto_images_ALL_SHEETS.xlsx

3) Run:
   python3 fill_images_in_xlsx_multi_bing.py

OUTPUT
------
products_with_images_ALL_SHEETS_BING.xlsx

TUNING
------
- OPTIONAL_QUERY_COLUMNS: add/remove columns to enrich the query.
- ALLOWED_DOMAINS: set to [] to allow any site; or e.g. ["pixabay.com","pexels.com"].
- DELAY_BETWEEN_QUERIES: throttle to be polite.
"""

import os
import time
import json
import urllib.parse
from typing import Optional, Dict, List
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ====================== USER CONFIG (EDIT THESE) ======================

INPUT_XLSX  = "products_for_auto_images_ALL_SHEETS.xlsx"
OUTPUT_XLSX = "products_with_images_ALL_SHEETS_BING.xlsx"

TITLE_HEADER = "Title"     # main query source
IMAGE_HEADER = "Image"     # where to write the found URL

# Use these columns (if present) to improve the search query
OPTIONAL_QUERY_COLUMNS = ["Brand", "Region"]

# If you want to keep images to “safer” sources, list domains here.
# Example: ["pixabay.com", "pexels.com", "unsplash.com"]   (licensing varies!)
# Leave [] to allow any domain.
ALLOWED_DOMAINS: List[str] = []

REQUEST_TIMEOUT = 20           # seconds for HTTP requests
DELAY_BETWEEN_QUERIES = 0.60   # seconds; be polite
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0 Safari/537.36"
)

# Try two pages of Bing Images if needed (first indices for page 1 and 2)
BING_PAGES_FIRST_PARAMS = [1, 11]   # 1..10, then 11..20

# ================== END OF USER CONFIG (USUALLY OK) ===================

HEADERS = {"User-Agent": USER_AGENT}


def head_ok(url: str, timeout: int = 10) -> bool:
    """Return True if the URL responds with 2xx/3xx (avoid dead links)."""
    try:
        r = requests.head(url, timeout=timeout, allow_redirects=True)
        return 200 <= r.status_code < 400
    except Exception:
        return False


def domain_is_allowed(url: str) -> bool:
    """If ALLOWED_DOMAINS is set, only accept URLs from those domains."""
    if not ALLOWED_DOMAINS:
        return True
    try:
        from urllib.parse import urlparse
        host = (urlparse(url).netloc or "").lower()
        return any(host.endswith(d.lower()) for d in ALLOWED_DOMAINS)
    except Exception:
        return False


def dedupe_words(text: str) -> str:
    """Remove duplicate words while preserving order (case-insensitive)."""
    seen = set()
    out = []
    for w in text.split():
        wl = w.lower()
        if wl in seen:
            continue
        seen.add(wl)
        out.append(w)
    return " ".join(out)


def build_query(base_title: str, ws, row_idx: int, col_index_by_name: Dict[str, int]) -> str:
    """
    Build a richer query: Title + optional Brand/Region (if present),
    then de-duplicate to avoid repeated words like "India India".
    """
    bits = [str(base_title).strip()]
    for col_name in OPTIONAL_QUERY_COLUMNS:
        idx = col_index_by_name.get(col_name)
        if idx:
            val = ws.cell(row=row_idx, column=idx).value
            if val and str(val).strip():
                bits.append(str(val).strip())

    # You can bias results by adding generic terms:
    # bits.append("gift card")  # Uncomment if it helps your data set.

    return dedupe_words(" ".join(bits))


def parse_bing_image_results(html: str) -> List[str]:
    """
    Extract candidate image URLs from a Bing Images HTML page.
    Strategy:
    - Prefer anchors <a class="iusc"> that carry a JSON-ish "m" attribute with "murl".
    - Fallback: <img> tags with http(s) src or data-src that looks like an image.
    Return a list of candidate URLs in the order they appear.
    """
    soup = BeautifulSoup(html, "lxml")
    candidates: List[str] = []

    # Primary: anchors with class iusc (metadata in 'm' attribute)
    for a in soup.select("a.iusc"):
        m_attr = a.get("m")
        if not m_attr:
            continue
        # 'm' is JSON-like; try to parse
        try:
            data = json.loads(m_attr)
            url = data.get("murl")
            if url and url.startswith("http"):
                candidates.append(url)
        except Exception:
            # sometimes it's just a string; do a naive extraction
            key = '"murl":"'
            m = m_attr
            if key in m:
                start = m.find(key) + len(key)
                end = m.find('"', start)
                if start > -1 and end > start:
                    url = m[start:end]
                    if url.startswith("http"):
                        candidates.append(url)

    # Fallback: any <img> with a plausible URL
    for img in soup.select("img"):
        src = img.get("data-src") or img.get("src") or ""
        if src.startswith("http") and any(ext in src.lower() for ext in [".jpg", ".jpeg", ".png", ".webp"]):
            candidates.append(src)

    # De-duplicate while preserving order
    seen = set()
    unique = []
    for u in candidates:
        if u in seen:
            continue
        seen.add(u)
        unique.append(u)
    return unique


def bing_first_live_image(query: str) -> Optional[str]:
    """
    Query Bing Images (page 1, then 2 if needed) and return the first
    URL that:
      - matches ALLOWED_DOMAINS (if configured),
      - and responds OK to a HEAD request.
    """
    q = urllib.parse.quote_plus(query)

    for first in BING_PAGES_FIRST_PARAMS:
        url = f"https://www.bing.com/images/search?q={q}&form=HDRSC2&first={first}&mkt=en-US"
        try:
            r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
        except Exception:
            continue

        for candidate in parse_bing_image_results(r.text):
            if not domain_is_allowed(candidate):
                continue
            if head_ok(candidate, timeout=10):
                return candidate

        # brief pause before trying next page
        time.sleep(0.3)

    return None


def ensure_headers(ws, headers_needed: List[str]) -> Dict[str, int]:
    """
    Ensure required headers exist in row 1. If missing, create them.
    Return a mapping {header_name: column_index}.
    """
    header_row = 1
    existing = {}
    max_col = ws.max_column or 1

    # read existing headers
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str) and val.strip():
            existing[val.strip()] = col

    # add missing headers to the right
    for h in headers_needed:
        if h not in existing:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=h)
            existing[h] = max_col

    return existing


def main():
    if not os.path.exists(INPUT_XLSX):
        print(f"[!] Input workbook not found: {INPUT_XLSX}")
        return

    wb = load_workbook(INPUT_XLSX)
    total_filled = 0
    total_skipped = 0

    for ws in wb.worksheets:
        col_map = ensure_headers(ws, [TITLE_HEADER, IMAGE_HEADER])
        title_col = col_map[TITLE_HEADER]
        image_col = col_map[IMAGE_HEADER]

        # also keep indices for optional query columns
        col_index_by_name = dict(col_map)

        rows = ws.max_row
        filled = 0
        skipped = 0

        print(f"\nSheet: {ws.title} — rows: {rows-1}")
        print(f"Filling '{IMAGE_HEADER}' from '{TITLE_HEADER}' using Bing Images...")

        for row in range(2, rows + 1):
            title_val = ws.cell(row=row, column=title_col).value

            # skip empty titles
            if not title_val or not str(title_val).strip():
                skipped += 1
                continue

            # skip if already has an image
            existing = ws.cell(row=row, column=image_col).value
            if existing and str(existing).strip():
                continue

            query = build_query(str(title_val), ws, row, col_index_by_name)
            print(f"- [{ws.title} R{row}] Bing image search: {query!r}")

            url = bing_first_live_image(query)

            if url:
                ws.cell(row=row, column=image_col, value=url)
                filled += 1
            else:
                print("  (no image found)")
            time.sleep(DELAY_BETWEEN_QUERIES)

        print(f"Sheet '{ws.title}': filled {filled}, skipped {skipped}")
        total_filled += filled
        total_skipped += skipped

    wb.save(OUTPUT_XLSX)
    print("\nAll done.")
    print(f"Total filled:  {total_filled}")
    print(f"Total skipped: {total_skipped}")
    print(f"Saved as:      {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
