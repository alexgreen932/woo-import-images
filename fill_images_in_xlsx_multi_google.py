#!/usr/bin/env python3
"""
fill_images_in_xlsx_multi_google.py
-----------------------------------

WHAT THIS DOES
--------------
- Opens your Excel workbook that contains MULTIPLE SHEETS.
- On every sheet, it looks for a column named "Title" (product title).
- It queries the Google Custom Search JSON API (CSE) in **image** mode.
- When it finds a result, it writes the image URL into the "Image" column.
- Saves a new workbook with the Image column filled.

WHY GOOGLE CSE?
---------------
- Official API (no fragile scraping).
- Better coverage than pure CC0 sources like Pixabay.
- You must have a Google API key and a CSE (cx) that searches the entire web
  and has "Image search" enabled.

REQUIREMENTS (one-time)
-----------------------
1) Python 3.12 + a virtual environment is recommended:

   python3 -m venv ~/woo-import-images/.venv
   source ~/woo-import-images/.venv/bin/activate
   pip install openpyxl requests

2) Google API key + Custom Search Engine:
   - Create API key: https://console.cloud.google.com/apis/credentials
   - Create Programmable Search Engine: https://programmablesearchengine.google.com/
     * Set it to search the **entire web**.
     * Enable **Image search** in the control panel.
     * Grab the CSE ID (cx).
   - Enable the **Custom Search API** for your project in Google Cloud.

3) Provide credentials to the script:
   EITHER export env vars:
     export GOOGLE_CSE_KEY='YOUR_API_KEY'
     export GOOGLE_CSE_CX='YOUR_CX_ID'
   OR pass CLI options:
     python3 fill_images_in_xlsx_multi_google.py --key 'YOUR_API_KEY' --cx 'YOUR_CX_ID'
   OR store in local files (first line only):
     echo 'YOUR_API_KEY' > google.key
     echo 'YOUR_CX_ID'   > google.cx

HOW TO RUN
----------
Place this script next to your workbook (the prepared one with Title + Image):
    products_for_auto_images_ALL_SHEETS.xlsx

Then run, for example:
    python3 fill_images_in_xlsx_multi_google.py --key 'YOUR_API_KEY' --cx 'YOUR_CX_ID'

OUTPUT
------
- A new file named:
    products_with_images_ALL_SHEETS_GOOGLE.xlsx

NOTES
-----
- Free quota is limited (usually ~100 queries/day). If you have thousands of rows,
  consider batching over days, adding billing, or fallbacks (Pixabay, Bing API).
- This script:
  * De-duplicates words in the query (so "India India Global" => "India Global").
  * Optionally uses extra columns (Brand, Region) to improve matches.
  * Skips rows where Image already exists.
  * HEAD-checks the chosen URL before writing it (avoid dead links).
"""

import os
import time
import argparse
from pathlib import Path
from typing import Optional, Dict, List

import requests
from openpyxl import load_workbook

# ====================== USER CONFIG (EDIT THESE) ======================

INPUT_XLSX  = "products_for_auto_images_ALL_SHEETS.xlsx"
OUTPUT_XLSX = "products_with_images_ALL_SHEETS_GOOGLE.xlsx"

TITLE_HEADER = "Title"     # Column to read the product title / main query
IMAGE_HEADER = "Image"     # Column to write the found image URL

# Use these optional columns to enrich the query (if they exist in the sheet)
OPTIONAL_QUERY_COLUMNS = ["Brand", "Region"]

REQUEST_TIMEOUT = 20       # seconds for HTTP requests
DELAY_BETWEEN_QUERIES = 0.45  # seconds; small delay to be polite
GOOGLE_ENDPOINT = "https://www.googleapis.com/customsearch/v1"

SAFE = "active"            # 'active'|'off' — safe search setting
IMG_SIZE = "large"         # 'icon','small','medium','large','xlarge','xxlarge','huge'
RESULTS_PER_QUERY = 10     # up to 10 per API call; we’ll pick the first live one

# ================== END OF USER CONFIG (USUALLY OK) ===================


# -------- Helpers to resolve credentials (key/cx) in multiple ways -----

def resolve_google_key_and_cx() -> (str, str):
    """
    Resolve GOOGLE_CSE_KEY and GOOGLE_CSE_CX in this order:
    1) CLI args (--key, --cx)
    2) Env vars (GOOGLE_CSE_KEY, GOOGLE_CSE_CX)
    3) Local files google.key, google.cx (first line only)
    4) Interactive prompt
    """
    parser = argparse.ArgumentParser(description="Fill Image column using Google CSE Image Search.")
    parser.add_argument("--key", help="Google API key (overrides env/file)")
    parser.add_argument("--cx",  help="Google CSE ID (overrides env/file)")
    args, _ = parser.parse_known_args()

    key = (args.key or os.getenv("GOOGLE_CSE_KEY", "")).strip()
    cx  = (args.cx  or os.getenv("GOOGLE_CSE_CX",  "")).strip()

    if not key:
        keyfile = Path("google.key")
        if keyfile.exists():
            key_txt = keyfile.read_text(encoding="utf-8").splitlines()
            if key_txt:
                key = key_txt[0].strip()

    if not cx:
        cxfile = Path("google.cx")
        if cxfile.exists():
            cx_txt = cxfile.read_text(encoding="utf-8").splitlines()
            if cx_txt:
                cx = cx_txt[0].strip()

    # last resort: ask
    if not key:
        try:
            key = input("Enter GOOGLE_CSE_KEY: ").strip()
        except KeyboardInterrupt:
            pass

    if not cx:
        try:
            cx = input("Enter GOOGLE_CSE_CX: ").strip()
        except KeyboardInterrupt:
            pass

    if not key or not cx:
        raise RuntimeError("Missing credentials. Provide --key and --cx, or set env vars, or files google.key/google.cx.")

    return key, cx


# ------------------------------ HTTP helpers ---------------------------

def head_ok(url: str, timeout: int = 10) -> bool:
    """Return True if URL responds 2xx/3xx; avoids storing dead links."""
    try:
        r = requests.head(url, timeout=timeout, allow_redirects=True)
        return 200 <= r.status_code < 400
    except Exception:
        return False


def google_image_search(key: str, cx: str, query: str, num: int = RESULTS_PER_QUERY) -> List[str]:
    """
    Query Google Custom Search JSON API in image mode and return a list of image URLs.
    Docs: https://developers.google.com/custom-search/v1/reference/rest/v1/cse/list
    """
    params = {
        "key": key,
        "cx": cx,
        "q": query,
        "searchType": "image",    # crucial: image mode
        "safe": SAFE,
        "imgSize": IMG_SIZE,
        "num": min(max(1, num), 10),  # Google allows up to 10 per request
    }
    try:
        r = requests.get(GOOGLE_ENDPOINT, params=params, timeout=REQUEST_TIMEOUT)
        # Simple handling for quota errors / 429 etc.
        if r.status_code == 429:
            # Too many requests — wait a bit more, user may need to add billing or slow down
            time.sleep(2.0)
            r = requests.get(GOOGLE_ENDPOINT, params=params, timeout=REQUEST_TIMEOUT)
        r.raise_for_status()
    except Exception:
        return []

    data = r.json()
    items = data.get("items", []) or []
    urls = []
    for it in items:
        link = it.get("link")  # direct image URL
        if link:
            urls.append(link)
    return urls


# -------------------------- Excel helpers ------------------------------

def ensure_headers(ws, headers_needed: List[str]) -> Dict[str, int]:
    """
    Ensure required headers exist in row 1. If a header is missing, create it.
    Return a mapping {header_name: column_index}.
    """
    header_row = 1
    existing = {}
    max_col = ws.max_column or 1

    # read current headers
    for col in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, str) and val.strip():
            existing[val.strip()] = col

    # add missing headers to the right
    for h in headers_needed:
        if h not in existing:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=h)
            existing[h] = max_col

    return existing


# ----------------------- Query building utilities ----------------------

def dedupe_words(text: str) -> str:
    """
    Remove duplicate words while preserving order.
    E.g., "Aeropostale India India Global" -> "Aeropostale India Global"
    """
    seen = set()
    parts = []
    for w in text.split():
        if w.lower() in seen:
            continue
        seen.add(w.lower())
        parts.append(w)
    return " ".join(parts)


def build_query(base_title: str, ws, row_idx: int, col_index_by_name: Dict[str, int]) -> str:
    """
    Construct a search query:
    - Start with Title.
    - Append optional columns (Brand, Region) if present.
    - De-duplicate words.
    You can customize this freely (e.g., append "gift card" or "prepaid voucher").
    """
    bits = [str(base_title).strip()]
    for col_name in OPTIONAL_QUERY_COLUMNS:
        idx = col_index_by_name.get(col_name)
        if idx:
            val = ws.cell(row=row_idx, column=idx).value
            if val and str(val).strip():
                bits.append(str(val).strip())

    # Optionally bias toward your domain:
    # bits.append("gift card")  # uncomment if helpful

    query = dedupe_words(" ".join(bits))
    return query


# --------------------------------- MAIN --------------------------------

def main():
    if not os.path.exists(INPUT_XLSX):
        print(f"[!] Input workbook not found: {INPUT_XLSX}")
        return

    key, cx = resolve_google_key_and_cx()

    wb = load_workbook(INPUT_XLSX)
    total_filled = 0
    total_skipped = 0

    for ws in wb.worksheets:
        col_map = ensure_headers(ws, [TITLE_HEADER, IMAGE_HEADER])
        title_col = col_map[TITLE_HEADER]
        image_col = col_map[IMAGE_HEADER]

        # Keep a mapping we can also use to read optional columns
        col_index_by_name = {k: v for k, v in col_map.items()}
        rows = ws.max_row
        filled = 0
        skipped = 0

        print(f"\nSheet: {ws.title} — rows: {rows-1}")
        print(f"Filling '{IMAGE_HEADER}' from '{TITLE_HEADER}' using Google CSE...")

        for row in range(2, rows + 1):
            title_val = ws.cell(row=row, column=title_col).value

            # 1) Skip rows without a title
            if not title_val or not str(title_val).strip():
                skipped += 1
                continue

            # 2) If Image already set, skip
            existing_image = ws.cell(row=row, column=image_col).value
            if existing_image and str(existing_image).strip():
                continue

            # 3) Make query (Title + optional columns, deduped)
            query = build_query(str(title_val), ws, row, col_index_by_name)
            print(f"- [{ws.title} R{row}] Google image search: {query!r}")

            # 4) Query Google CSE
            candidates = google_image_search(key, cx, query, num=RESULTS_PER_QUERY)

            # 5) Pick the first live URL
            chosen = None
            for url in candidates:
                if head_ok(url, timeout=10):
                    chosen = url
                    break

            if chosen:
                ws.cell(row=row, column=image_col, value=chosen)
                filled += 1
            else:
                print("  (no live image found)")

            time.sleep(DELAY_BETWEEN_QUERIES)

        print(f"Sheet '{ws.title}': filled {filled}, skipped {skipped}")
        total_filled += filled
        total_skipped += skipped

    wb.save(OUTPUT_XLSX)
    print("\nAll done.")
    print(f"Total filled:  {total_filled}")
    print(f"Total skipped: {total_skipped}")
    print(f"Saved as:      {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
